import csv
import re
import unicodedata
from datetime import date as date_type, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from comercial.models import Proposta, PropostaRevisao
from financeiro.models import Empresa
from pessoas.models import Pessoa


def normalizar(texto):
    texto = unicodedata.normalize("NFKD", str(texto or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]", "", texto.upper())


def moeda(valor):
    texto = str(valor or "0").strip().replace("R$", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    return Decimal(texto or "0")


def data(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date_type):
        return valor
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(str(valor).strip(), formato).date()
        except ValueError:
            pass
    raise ValueError("data inválida")


def normalizar_cabecalho(texto):
    texto = unicodedata.normalize("NFKD", str(texto or "")).encode("ascii", "ignore").decode().lower().strip()
    return re.sub(r"[^a-z0-9]+", "", texto)


ALIASES_COLUNAS = {
    "numero": {"numero", "numeroproposta", "noprop", "nprop", "nproposta", "proposta"},
    "data": {"data", "dataemissao", "datadeemissao"},
    "emitente": {"emitente", "responsavel", "responsavelinterno"},
    "cliente": {"cliente"},
    "contato": {"contato", "aoscuidados", "responsavelcliente"},
    "descricao": {"descricao", "servico", "nomeservico"},
    "status": {"status", "statushistorico"},
    "valor": {"valor", "valortotal", "total"},
    "observacao": {"observacao", "observacoes", "informacaocomplementar", "complemento"},
}
COLUNAS_OBRIGATORIAS = {"numero", "data", "cliente", "descricao", "status", "valor"}
CLIENTES_CONFIRMADOS = {
    "MDZ": {"razao_social": "MONDELEZ BRASIL NORTE NORDESTE LTDA", "cpf_cnpj": "10144076000144", "cidade": "Vitória de Santo Antão", "estado": "PE", "referencia": "MDZ / MDLZ"},
    "MDLZ": {"razao_social": "MONDELEZ BRASIL NORTE NORDESTE LTDA", "cpf_cnpj": "10144076000144", "cidade": "Vitória de Santo Antão", "estado": "PE", "referencia": "MDZ / MDLZ"},
    "FQM": {"razao_social": "FARMOQUIMICA S A", "cpf_cnpj": "33349473001634", "cidade": "Pombos", "estado": "PE", "referencia": "FQM"},
    "CAPRICCHE": {"razao_social": "CAPRICCHE S.A.", "cpf_cnpj": "17090600000190", "cidade": "Moreno", "estado": "PE", "referencia": "CAPRICCHE"},
    "DOCILE": {"razao_social": "DOCILE NORDESTE - INDUSTRIA E COMERCIO DE PRODUTOS ALIMENTICIOS LTDA", "cpf_cnpj": "12020480000131", "cidade": "Vitória de Santo Antão", "estado": "PE", "referencia": "DOCILE"},
    "MASTERBOITO": {"razao_social": "MASTERBOI LTDA.", "cpf_cnpj": "03721769000600", "cidade": "Nova Olinda", "estado": "TO", "referencia": "MASTERBOI/TO"},
}
CLIENTES_INCOMPLETOS = {"ARB", "MJENGENHARIA", "VEOLIA"}


def dados_cliente_historico(nome):
    chave = normalizar(nome)
    if chave in CLIENTES_CONFIRMADOS:
        return {**CLIENTES_CONFIRMADOS[chave], "incompleto": False}
    return {
        "razao_social": str(nome or "").strip(), "cpf_cnpj": "", "cidade": "", "estado": "",
        "referencia": str(nome or "").strip(), "incompleto": chave in CLIENTES_INCOMPLETOS,
    }


def interpretar_codigo_historico(valor):
    codigo = "".join(str(valor or "").upper().split())
    encontrado = re.fullmatch(r"(VERS\d+)(?:\.(\d+))?", codigo)
    if not encontrado:
        raise ValueError("número fora do padrão histórico VERS")
    revisao = int(encontrado.group(2) or 0)
    if encontrado.group(2) is not None and revisao <= 0:
        raise ValueError("número da revisão deve ser inteiro positivo")
    return encontrado.group(1), revisao


def mapear_cabecalho(valores):
    aliases = {alias: destino for destino, opcoes in ALIASES_COLUNAS.items() for alias in opcoes}
    mapeamento = {}
    for indice, valor in enumerate(valores):
        destino = aliases.get(normalizar_cabecalho(valor))
        if destino and destino not in mapeamento:
            mapeamento[destino] = indice
    # Na planilha real, H2 contém =SUBTOTAL(...) e o valor cacheado é zero.
    # Aceitamos esse caso somente na coluna imediatamente após Status.
    if "valor" not in mapeamento and "status" in mapeamento:
        indice_valor = mapeamento["status"] + 1
        if indice_valor < len(valores) and valores[indice_valor] == 0:
            mapeamento["valor"] = indice_valor
    return mapeamento


class Command(BaseCommand):
    TAMANHO_MAXIMO = 20 * 1024 * 1024
    LINHAS_MAXIMAS = 10000
    help = "Analisa/importa propostas históricas de CSV ou XLSX de forma idempotente."

    def add_arguments(self, parser):
        parser.add_argument("arquivo")
        parser.add_argument("--empresa", type=int)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--data-inicial")

    def _converter_linhas(self, linhas):
        linhas = iter(linhas)
        cabecalho = None
        for valores in linhas:
            mapeamento = mapear_cabecalho(valores)
            if COLUNAS_OBRIGATORIAS.issubset(mapeamento):
                cabecalho = mapeamento
                break
        if cabecalho is None:
            raise CommandError("Cabeçalho não encontrado ou sem as colunas obrigatórias: número, data, cliente, descrição, status e valor.")
        for valores in linhas:
            if not any(valor not in (None, "") for valor in valores):
                continue
            yield {campo: valores[indice] if indice < len(valores) else None for campo, indice in cabecalho.items()}

    def _linhas(self, caminho):
        if caminho.suffix.lower() == ".csv":
            with caminho.open(encoding="utf-8-sig", newline="") as arquivo:
                primeira = arquivo.readline()
                arquivo.seek(0)
                yield from self._converter_linhas(csv.reader(arquivo, delimiter=";" if ";" in primeira else ","))
            return
        if caminho.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise CommandError("Instale openpyxl para importar XLSX.") from exc
            workbook = load_workbook(caminho, read_only=True, data_only=True)
            try:
                planilha = workbook.active
                yield from self._converter_linhas(planilha.iter_rows(min_col=1, max_col=20, values_only=True))
            finally:
                workbook.close()
            return
        raise CommandError("Use um arquivo .csv ou .xlsx.")

    def handle(self, *args, **opcoes):
        caminho = Path(opcoes["arquivo"])
        if not caminho.exists():
            raise CommandError("Arquivo não encontrado.")
        if caminho.stat().st_size > self.TAMANHO_MAXIMO:
            raise CommandError("Arquivo excede o limite de 20 MB.")
        data_inicial = None
        if opcoes.get("data_inicial"):
            try:
                data_inicial = datetime.strptime(opcoes["data_inicial"], "%Y-%m-%d").date()
            except ValueError as exc:
                raise CommandError("--data-inicial deve usar o formato ISO YYYY-MM-DD.") from exc
        empresas = Empresa.objects.filter(pk=opcoes.get("empresa")) if opcoes.get("empresa") else Empresa.objects.filter(ativa=True)
        if empresas.count() != 1:
            raise CommandError("Informe --empresa quando houver zero ou mais de uma empresa ativa.")
        empresa = empresas.get()
        relatorio = {
            "linhas_recorte": 0, "propostas_base": 0, "revisoes_historicas": 0,
            "validas": 0, "duplicadas": 0, "erros_estruturais": 0,
            "datas_invalidas_ignoradas": 0, "anteriores_corte": 0,
            "clientes_novos": 0, "clientes_incompletos": 0,
            "revisoes_sem_base": 0, "alertas_valor_zero": 0,
            "propostas_importadas": 0, "revisoes_importadas": 0,
        }
        linhas = list(self._linhas(caminho))
        if len(linhas) > self.LINHAS_MAXIMAS:
            raise CommandError(f"Arquivo excede o limite de {self.LINHAS_MAXIMAS} registros.")
        bases_arquivo = set()
        for linha in linhas:
            try:
                base, revisao = interpretar_codigo_historico(linha.get("numero"))
                if revisao == 0:
                    bases_arquivo.add(base)
            except ValueError:
                pass
        preparados = []
        chaves_arquivo = set()
        bases_validas = set()
        clientes_novos = set()
        existentes = {}
        existentes_documento = {}
        for pessoa in Pessoa.objects.all():
            existentes.setdefault(normalizar(pessoa.razao_social), []).append(pessoa)
            documento = re.sub(r"\D", "", pessoa.cpf_cnpj or "")
            if documento:
                existentes_documento.setdefault(documento, []).append(pessoa)
        recorte_iniciado = False
        for numero, linha in enumerate(linhas, start=2):
            try:
                try:
                    data_proposta = data(linha.get("data"))
                except ValueError:
                    if data_inicial and not recorte_iniciado:
                        relatorio["datas_invalidas_ignoradas"] += 1
                        continue
                    raise
                if data_inicial and data_proposta < data_inicial:
                    relatorio["anteriores_corte"] += 1
                    continue
                recorte_iniciado = True
                relatorio["linhas_recorte"] += 1
                codigo, revisao = interpretar_codigo_historico(linha.get("numero"))
                proposta_existente = Proposta.objects.filter(empresa=empresa, codigo=codigo).first()
                if revisao and codigo not in bases_arquivo and not proposta_existente:
                    relatorio["revisoes_sem_base"] += 1
                    self.stderr.write(f"Linha {numero}: revisão {revisao:02d} bloqueada; proposta-base {codigo} ausente")
                    continue
                chave = (codigo, revisao)
                if proposta_existente and proposta_existente.revisoes.filter(numero=revisao).exists():
                    relatorio["duplicadas"] += 1
                    continue
                if chave in chaves_arquivo:
                    relatorio["duplicadas"] += 1
                    continue
                chaves_arquivo.add(chave)
                bases_validas.add(codigo)
                cliente_nome = str(linha.get("cliente") or "").strip()
                if not cliente_nome:
                    raise ValueError("cliente ausente")
                dados_cliente = dados_cliente_historico(cliente_nome)
                valor = moeda(linha.get("valor"))
                relatorio["validas"] += 1
                if revisao:
                    relatorio["revisoes_historicas"] += 1
                if valor == 0:
                    relatorio["alertas_valor_zero"] += 1
                if dados_cliente["incompleto"]:
                    relatorio["clientes_incompletos"] += 1
                candidatos = existentes_documento.get(dados_cliente["cpf_cnpj"], []) if dados_cliente["cpf_cnpj"] else existentes.get(normalizar(dados_cliente["razao_social"]), [])
                if len(candidatos) > 1:
                    raise ValueError("mais de um cadastro equivalente para o cliente")
                cliente = candidatos[0] if candidatos else None
                chave_cliente = dados_cliente["cpf_cnpj"] or normalizar(dados_cliente["razao_social"])
                if not cliente and chave_cliente not in clientes_novos:
                    relatorio["clientes_novos"] += 1
                    clientes_novos.add(chave_cliente)
                preparados.append({
                    "codigo": codigo, "revisao": revisao, "cliente": cliente,
                    "cliente_nome": dados_cliente["razao_social"], "cliente_historico": cliente_nome,
                    "cliente_dados": dados_cliente, "data": data_proposta,
                    "servico": str(linha.get("descricao") or "Proposta histórica").strip(),
                    "contato": str(linha.get("contato") or "").strip(), "valor": valor,
                    "status_historico": str(linha.get("status") or "").strip(),
                    "emitente": str(linha.get("emitente") or "").strip(),
                    "observacao": str(linha.get("observacao") or "").strip(),
                })
            except (ValueError, InvalidOperation) as erro:
                relatorio["erros_estruturais"] += 1
                self.stderr.write(f"Linha {numero}: {erro}")
        relatorio["propostas_base"] = len(bases_validas)
        if not opcoes["dry_run"]:
            with transaction.atomic():
                clientes_criados = {}
                for item in preparados:
                    dados_cliente = item["cliente_dados"]
                    chave_cliente = dados_cliente["cpf_cnpj"] or normalizar(item["cliente_nome"])
                    cliente = item["cliente"] or clientes_criados.get(chave_cliente)
                    if not cliente:
                        observacoes = f"Importação histórica: identificado na planilha como {dados_cliente['referencia']}."
                        if dados_cliente["incompleto"]:
                            observacoes += " Cadastro incompleto — revisar posteriormente."
                        cliente = Pessoa.objects.create(
                            razao_social=item["cliente_nome"], cpf_cnpj=dados_cliente["cpf_cnpj"],
                            cidade=dados_cliente["cidade"], estado=dados_cliente["estado"],
                            observacoes=observacoes, classificacao=Pessoa.Classificacao.CLIENTE,
                        )
                        clientes_criados[chave_cliente] = cliente
                    proposta, criada = Proposta.objects.get_or_create(
                        empresa=empresa, codigo=item["codigo"],
                        defaults={"cliente": cliente, "numero_sequencial": int(item["codigo"][4:]), "origem": Proposta.Origem.IMPORTADO_HISTORICO},
                    )
                    notas = f"Status histórico: {item['status_historico']}\nEmitente: {item['emitente']}"
                    if item["observacao"]:
                        notas += f"\nObservação histórica: {item['observacao']}"
                    PropostaRevisao.objects.create(
                        proposta=proposta, numero=item["revisao"], data_proposta=item["data"],
                        nome_servico=item["servico"], aos_cuidados_de=item["contato"],
                        formacao_preco=PropostaRevisao.FormacaoPreco.MANUAL,
                        preco_venda_final=item["valor"], observacoes_internas=notas, congelada=True,
                    )
                    proposta.revisao_atual = max(proposta.revisao_atual, item["revisao"])
                    proposta.status_historico = item["status_historico"]
                    proposta.observacao_importacao = item["observacao"]
                    proposta.save(update_fields=["revisao_atual", "status_historico", "observacao_importacao", "atualizado_em"])
                    relatorio["propostas_importadas"] += int(criada)
                    relatorio["revisoes_importadas"] += 1
        modo = "DRY-RUN" if opcoes["dry_run"] else "IMPORTAÇÃO"
        self.stdout.write(self.style.SUCCESS(f"{modo}: {relatorio}"))
